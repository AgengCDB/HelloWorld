import pandas as pd
import os
import glob
from datetime import datetime
import re
import openpyxl
import shutil
import aspose.pdf as ap  # Ensure you have installed the aspose-pdf package
import warnings
from datetime import datetime

# Get the current time
current_time = datetime.now()

# Suppress FutureWarning related to treating keys as positions in pandas
warnings.filterwarnings("ignore", category=FutureWarning)

def getRenewal(input_folder="", output_folder="DEFAULT", converted_folder="DEFAULT", is_fileName=False, is_moveConverted=False):
    """ 
    Function that convert PDF File into Excel

    Args:
        input_folder (_type_): use this folder to input your pdf(s)
        output_folder (_type_): all the xlsx (from converted pdf) will be here
        converted_folder (_type_): pdf will be moved here if conversion success
    """
    if not input_folder or not os.listdir(input_folder):
        e = (f"{current_time}\tError, Input folder is empty")
        return e
    if output_folder == "DEFAULT":
        output_folder = input_folder+"/output/"
        os.makedirs(output_folder, exist_ok=True)
    if converted_folder == "DEFAULT":
        converted_folder = input_folder+"/converted/"
        os.makedirs(converted_folder, exist_ok=True)
        
    # Ensure the output and converted directories exist
    print("========== START PDF TO XLSX ==========")
    
    os.makedirs(converted_folder, exist_ok=True)
    i = 1

    # Get all PDF files in the input directory and its subdirectories
    for root, dirs, files in os.walk(input_folder):
        for file_name in files:
            if file_name.endswith(".pdf"):
                pdf_file = os.path.join(root, file_name)
                print("===== "+str(i))
                try:
                    # Construct the output file path with .xlsx extension
                    base_name = os.path.basename(pdf_file)
                    output_file = os.path.join(output_folder, os.path.splitext(base_name)[0] + ".xlsx")

                    # Open PDF document
                    document = ap.Document(pdf_file)

                    # Define save options for Excel format
                    save_option = ap.ExcelSaveOptions()

                    # Save the file into MS Excel format
                    document.save(output_file, save_option)
                    print(f"Converted {pdf_file} to {output_file}")

                    workbook = openpyxl.load_workbook(output_file)
                    for sheet_name in workbook.sheetnames:
                        worksheet = workbook[sheet_name]
                        for row in worksheet.iter_rows():
                            for cell in row:
                                cell.number_format = '@'
                    workbook.save(output_file)

                    if is_moveConverted:
                        # Move the original PDF to the converted folder
                        shutil.move(pdf_file, os.path.join(converted_folder, base_name))
                        print(f"Moved {pdf_file} to {os.path.join(converted_folder, base_name)}")
                        
                except Exception as e:
                    # Handle conversion errors
                    print(f"Failed to convert {pdf_file}: {e}")
                i += 1
    print("========== FINISH PDF TO XLSX ==========\n")

    ### Output as excel
    now = datetime.now()
    time_string = now.strftime('%Y%m%d_%H%M%S')
    df_result = process_all_excels_in_folder(output_folder)
    # Clean currency
    df_result['NILAI BHP'] = df_result['NILAI BHP'].apply(process_nilai_bhp)
    
    if not is_fileName: df_result.drop(columns="fileName", inplace=True)

    df_result.to_excel(f"{input_folder}\\Renewal_{time_string}.xlsx", index=False)
    print(f"Renewal saved as \n{input_folder}\\Renewal_{time_string}.xlsx")
    return f"{input_folder}\\Renewal_{time_string}.xlsx"

def process_all_excels_in_folder(input_folder):
    # Get all Excel files in the input directory
    excel_files = glob.glob(os.path.join(input_folder, "*.xlsx"))

    # Initialize an empty DataFrame to store all results
    all_results_df = pd.DataFrame()

    # Process each Excel file
    for excel_file in excel_files:
        result_df = extract_data_from_excel(excel_file)
        all_results_df = pd.concat([all_results_df, result_df], ignore_index=True)

    return all_results_df

def process_nilai_bhp(value):
    # Remove the "Rp" prefix and ",00" suffix
    value = value.replace('Rp', '').replace(',00', '')
    # Remove all "." from the string
    value = value.replace('.', '')
    # Convert the result to a number
    number = int(value)
    # Convert the number back to a text string
    return number

def extract_data_from_excel(file_path, sheet_name='Sheet2'):
    df = pd.read_excel(file_path, sheet_name)

    temp_column = ['fileName', 'Nomor Aplikasi', 'SITE ID', 'NILAI BHP']
    dfHasil = pd.DataFrame(columns=temp_column)

    # Initialize variables
    nomor_aplikasi = None

    # Assuming df is your DataFrame already loaded from the Excel file
    for i, row in df.iterrows():
        if isinstance(row[0], str) and 'nomor aplikasi' in row[0].lower():
            nomor_aplikasi = row[1]
            continue     
        elif isinstance(row[0], str) and 'SITE ID' in row[0]:
            j = i + 1  # Start checking from the next row
            while j < len(df):
                next_row = df.iloc[j]
                if not isinstance(next_row[0], int):
                    j += 1 ### skip row
                    continue
                else:
                    ### GET SITE ID
                    site_id = next_row[0]
                    
                    # Extract currency value from the next row's columns
                    currency_values = []
                    for cell_value in next_row[1:]:  # Start from the 2nd column
                        if isinstance(cell_value, str):
                            currency_value = extract_currency_value(cell_value)
                            if currency_value is not None:
                                currency_values.append(currency_value)
                    
                    nilai_bhp = ', '.join(map(str, currency_values))
                    new_row = {'fileName': file_path, 'Nomor Aplikasi': nomor_aplikasi, 'SITE ID': site_id, 'NILAI BHP': nilai_bhp}
                    
                    # Insert to dataframe
                    dfHasil = pd.concat([dfHasil, pd.DataFrame([new_row])], ignore_index=True)
                    
                    j += 1
            continue
    return dfHasil

def extract_currency_value(text):
    # Define a regular expression pattern to match the currency value
    pattern = r'Rp([\d.,]+),00'
    match = re.search(pattern, text)
    if match:
        # Extract the matched currency value
        currency_value = match.group(1)
        # Remove non-numeric characters
        numeric_value = re.sub(r'[^\d]', '', currency_value)
        return numeric_value
    else:
        return None
    
def delete_all_files_in_folder(folder_path):
    # Get a list of all files in the folder
    files = os.listdir(folder_path)

    # Iterate over each file and delete it
    for file_name in files:
        file_path = os.path.join(folder_path, file_name)
        if os.path.isfile(file_path):
            try:
                os.remove(file_path)
                print(f"Deleted file: {file_path}")
            except Exception as e:
                print(f"Failed to delete file: {file_path}, Error: {e}")